"""
NBA Studio — Tangible Lab wrapper.

Avvio:
    python tangible_lab/server.py

Questo modulo NON modifica il backend del cliente: importa l'app FastAPI di
nba_api come-è e aggiunge:
- un mount /lab che serve i file statici della Lab
- endpoint custom /lab/api/* (auth, casi, review, commenti) su SQLite
- endpoint /lab/admin/* protetti (richiedono role=admin)
- redirect /lab → /lab/index.html (se loggato) o /lab/login.html

URL Lab: http://127.0.0.1:<porta>/lab/
"""

from __future__ import annotations

import csv
import io
import json
import os
import socket
import sys
import threading
import webbrowser
from typing import Optional

# Assicura che il backend del cliente (nba_api.py in repo root) sia importabile
_REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

import shutil  # noqa: E402

# ----------- Seed dei file mutevoli su /data (solo se DATA_DIR è settato) -----------
# Sposta `dataset.json`, `nba_config.json`, `trigger_catalog_overrides.json` su volume
# persistente per non perderne le modifiche tra restart del container. I default
# vengono copiati una sola volta (se i file non esistono ancora su /data).
_DATA_DIR = os.environ.get("TANGIBLE_LAB_DATA_DIR")
if _DATA_DIR:
    os.makedirs(_DATA_DIR, exist_ok=True)
    def _seed(filename: str) -> str:
        target = os.path.join(_DATA_DIR, filename)
        src = os.path.join(_REPO_ROOT, filename)
        if not os.path.exists(target) and os.path.exists(src):
            shutil.copy2(src, target)
        return target
    _seeded_dataset = _seed("dataset.json")
    _seeded_config = _seed("nba_config.json")
    _seeded_overrides = _seed("trigger_catalog_overrides.json")

from nba_api import app  # noqa: E402 — import dopo aver fissato sys.path

# Dopo l'import, ridirigiamo i path globali a /data (lookup dinamico, le funzioni
# nei moduli leggono la variabile al momento dell'uso).
if _DATA_DIR:
    import nba_api as _nba_api
    import nba_config as _nba_config
    import nba_catalog as _nba_catalog
    _nba_api.DATASET_PATH = _seeded_dataset
    _nba_config.CONFIG_JSON_PATH = _seeded_config
    # nba_api ha fatto `from nba_config import CONFIG_JSON_PATH` (cattura per valore):
    # va riassegnato anche lì, altrimenti GET/PUT /config leggono/scrivono il path di
    # default (repo) invece di quello seedato → la config persiste nel posto sbagliato.
    _nba_api.CONFIG_JSON_PATH = _seeded_config
    _nba_catalog.OVERRIDES_PATH = _seeded_overrides
    # nba_config carica e CACHA `CONFIG` all'import (avvenuto sopra con
    # `from nba_api import app`), quando CONFIG_JSON_PATH puntava ancora al default.
    # Senza questo reload, get_config() resta sulla config vecchia/di default e lo
    # scoring (es. opportunity via avg_premiums) usa valori errati. Ricarichiamo dal
    # path seedato appena riassegnato.
    _nba_config.reload_config()

from fastapi import Body, HTTPException, Request, Response  # noqa: E402
from fastapi.responses import RedirectResponse, StreamingResponse  # noqa: E402
from fastapi.staticfiles import StaticFiles  # noqa: E402

# Importiamo le funzioni interne dell'engine cliente per esporre un breakdown
# leggibile del punteggio, senza modificare nba_engine.py.
from nba_engine import (  # noqa: E402
    client_from_dict, lead_from_dict,
    detect_client_triggers, detect_lead_triggers,
    client_urgency_score, client_value_score, client_opportunity_score, client_recency_score,
    client_relationship_modifier, churn_override,
)
# NB: client_multi_trigger_bonus rimosso dal motore nella v2 (20/06/2026): il punteggio
# ora è _clamp(base + relationship_modifier), senza bonus multi-trigger.
from nba_config import get_config, temporary_config  # noqa: E402

# Tangible Lab modules
from tangible_lab.db import init_db  # noqa: E402
from tangible_lab.auth import (  # noqa: E402
    hash_password, verify_password,
    create_session, delete_session,
    user_from_request, require_user, require_admin,
    set_session_cookie, clear_session_cookie, COOKIE_NAME,
)
from tangible_lab import models  # noqa: E402
from tangible_lab import checkup_engine, checkup_models  # noqa: E402
from tangible_lab import messages as lab_messages  # noqa: E402


LAB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")

# Init DB (idempotente)
init_db()


def _seed_checkup_cases_if_empty():
    """Al primo boot, popola checkup_cases con scenari demo (condivisi col team).
    Idempotente: salta se ci sono già righe nella tabella."""
    seed_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "checkup", "seed_cases.json")
    if not os.path.exists(seed_path):
        return
    from tangible_lab.db import get_conn
    from tangible_lab import models, checkup_engine, checkup_models
    import json as _json
    with get_conn() as conn:
        existing = conn.execute("SELECT COUNT(*) AS c FROM checkup_cases").fetchone()["c"]
    if existing > 0:
        return
    # Trova l'utente admin (bootstrap) come owner
    admins = [u for u in models.list_users() if u["role"] == "admin"]
    if not admins:
        return
    owner_id = admins[0]["id"]
    try:
        scenarios = _json.loads(open(seed_path, "r", encoding="utf-8").read())
        for s in scenarios:
            answers = s.get("answers") or {}
            errs = checkup_engine.validate_answers(answers)
            if errs:
                print(f"[tangible_lab] Skip seed '{s.get('name')}': {errs}")
                continue
            result = checkup_engine.compute(answers)
            checkup_models.create(owner_id, s["name"], answers, result, shared=True, notes="Scenario demo")
        print(f"[tangible_lab] Seed: caricati {len(scenarios)} scenari demo nel Check-up (condivisi col team)")
    except Exception as e:
        print(f"[tangible_lab] Errore seed checkup: {e}")


_seed_checkup_cases_if_empty()

# ============================== redirect everything-not-lab → /lab/ ==============================
# Nasconde l'interfaccia originale del cliente (home.html, NBA_Index.html, ecc.)
# e forza tutti gli URL "non-Lab" a passare per il login.
# Lasciamo accessibili solo:
#   - /lab/*           (frontend NBA Studio + API custom Tangible)
#   - /nba/*           (API engine — chiamate dal frontend Lab)
#   - /config, /config/*   (config engine — letta/scritta dal Lab)
#   - /catalog/*       (trigger catalog — letto/scritto dal Lab)
#   - /dataset/*       (reload dataset — usato dal Lab)
#   - /favicon.ico     (per non sporcare di redirect i log)
_LAB_ALLOWED_EXACT = {"/lab", "/config", "/favicon.ico"}
_LAB_ALLOWED_PREFIXES = ("/lab/", "/nba/", "/config/", "/catalog/", "/dataset/")


@app.middleware("http")
async def _redirect_non_lab_to_lab(request: Request, call_next):
    path = request.url.path
    if path in _LAB_ALLOWED_EXACT or path.startswith(_LAB_ALLOWED_PREFIXES):
        return await call_next(request)
    # UI originale Vittoria (home + /static/*) accessibile SOLO agli utenti loggati.
    # Non loggati → redirect a /lab/login.html
    if path == "/" or path.startswith("/static/"):
        user = user_from_request(request)
        if user:
            return await call_next(request)
        return RedirectResponse(url="/lab/login.html", status_code=307)
    return RedirectResponse(url="/lab/", status_code=307)


@app.get("/lab", include_in_schema=False)
def _lab_root_redirect(request: Request):
    user = user_from_request(request)
    return RedirectResponse(url="/lab/" if user else "/lab/login.html")


# Endpoint /lab/index.html è statico — protezione applicata via JS
# (controllo /lab/api/me all'avvio; se 401 → redirect a login.html).
# Per l'admin.html applichiamo invece controllo lato statico-route esplicito:
@app.get("/lab/admin.html", include_in_schema=False)
def _admin_html(request: Request):
    user = user_from_request(request)
    if not user:
        return RedirectResponse(url="/lab/login.html")
    if user["role"] != "admin":
        return RedirectResponse(url="/lab/")
    # File statico, FastAPI lo serve via il mount; qui solo controllo
    return Response(
        open(os.path.join(LAB_DIR, "admin.html"), "rb").read(),
        media_type="text/html; charset=utf-8",
    )


# ============================== health (per Render/uptime) ==============================

@app.get("/lab/api/health", include_in_schema=False)
def lab_health():
    return {"status": "ok"}


# ============================== auth API ==============================

@app.post("/lab/api/login", include_in_schema=False)
def lab_login(payload: dict = Body(...)):
    username = (payload.get("username") or "").strip()
    password = payload.get("password") or ""
    if not username or not password:
        raise HTTPException(status_code=400, detail="username e password sono obbligatori")
    u = models.get_user_by_username(username)
    if not u or not u["active"] or not verify_password(password, u["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    signed = create_session(u["id"])
    resp = Response(content='{"ok":true}', media_type="application/json")
    set_session_cookie(resp, signed)
    return resp


@app.post("/lab/api/logout", include_in_schema=False)
def lab_logout(request: Request):
    signed = request.cookies.get(COOKIE_NAME)
    if signed:
        delete_session(signed)
    resp = Response(content='{"ok":true}', media_type="application/json")
    clear_session_cookie(resp)
    return resp


@app.get("/lab/api/me", include_in_schema=False)
def lab_me(request: Request):
    u = user_from_request(request)
    if not u:
        raise HTTPException(status_code=401, detail="Non autenticato")
    return u


@app.post("/lab/api/me/acknowledge-password", include_in_schema=False)
def lab_dismiss_pw_warning(request: Request):
    """L'utente corrente conferma di aver visto l'avviso 'devi cambiare password'.
    Disattiva il flag (utile quando si è già cambiata da admin panel)."""
    u = require_user(request)
    models.update_user(u["id"], must_change_password=False)
    return {"ok": True}


@app.post("/lab/api/me/password", include_in_schema=False)
def lab_change_password(request: Request, payload: dict = Body(...)):
    u = require_user(request)
    current = payload.get("current") or ""
    newpw = payload.get("new") or ""
    if len(newpw) < 6:
        raise HTTPException(status_code=400, detail="La nuova password deve avere almeno 6 caratteri")
    fullu = models.get_user_by_username(u["username"])
    if not fullu or not verify_password(current, fullu["password_hash"]):
        raise HTTPException(status_code=400, detail="Password corrente errata")
    models.update_user(u["id"], password_hash=hash_password(newpw), must_change_password=False)
    return {"ok": True}


# ============================== cases API ==============================

@app.get("/lab/api/cases", include_in_schema=False)
def api_list_cases(request: Request):
    u = require_user(request)
    return models.list_cases_visible_to(u["id"])


@app.post("/lab/api/cases", include_in_schema=False)
def api_create_case(request: Request, payload: dict = Body(...)):
    u = require_user(request)
    name = (payload.get("name") or "").strip() or "(senza nome)"
    type_ = payload.get("type")
    record = payload.get("record") or {}
    shared = bool(payload.get("shared", False))
    notes = payload.get("notes") or ""
    if type_ not in ("client", "lead"):
        raise HTTPException(status_code=400, detail="type deve essere client o lead")
    cid = models.create_case(u["id"], name, type_, record, shared, notes)
    return models.get_case(cid)


@app.put("/lab/api/cases/{case_id}", include_in_schema=False)
def api_update_case(case_id: int, request: Request, payload: dict = Body(...)):
    u = require_user(request)
    c = models.get_case(case_id)
    if not c:
        raise HTTPException(status_code=404, detail="Caso non trovato")
    if c["owner_id"] != u["id"] and u["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo l'autore (o admin) può modificare")
    models.update_case(
        case_id,
        name=payload.get("name"),
        type_=payload.get("type"),
        record=payload.get("record"),
        shared=payload.get("shared"),
        notes=payload.get("notes"),
    )
    return models.get_case(case_id)


@app.delete("/lab/api/cases/{case_id}", include_in_schema=False)
def api_delete_case(case_id: int, request: Request):
    u = require_user(request)
    c = models.get_case(case_id)
    if not c:
        raise HTTPException(status_code=404, detail="Caso non trovato")
    if c["owner_id"] != u["id"] and u["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo l'autore (o admin) può eliminare")
    models.delete_case(case_id)
    return {"ok": True}


# ============================== reviews + comments API ==============================

@app.get("/lab/api/reviews", include_in_schema=False)
def api_list_my_reviews(request: Request):
    """Tutte le review dell'utente corrente (per popolare i badge della lista)."""
    u = require_user(request)
    return models.list_all_user_reviews(u["id"])


@app.get("/lab/api/reviews/{target_key:path}", include_in_schema=False)
def api_reviews_for_target(target_key: str, request: Request):
    """Per uno specifico target: la mia review + tutte le review (per visione condivisa) + i commenti."""
    u = require_user(request)
    return {
        "mine": models.get_user_review(u["id"], target_key),
        "all": models.list_reviews_for_target(target_key),
        "comments": models.list_comments_for_target(target_key),
    }


@app.put("/lab/api/reviews/{target_key:path}", include_in_schema=False)
def api_set_review(target_key: str, request: Request, payload: dict = Body(...)):
    u = require_user(request)
    judgement = payload.get("judgement")
    if judgement is None:
        models.delete_review(u["id"], target_key)
        return {"ok": True}
    if judgement not in ("ok", "ko", "unsure"):
        raise HTTPException(status_code=400, detail="judgement deve essere ok|ko|unsure")
    reason = (payload.get("reason") or "").strip() or None
    reason_text = (payload.get("reason_text") or "").strip() or None
    # La motivazione è obbligatoria per ko e unsure
    if judgement in ("ko", "unsure"):
        if not reason:
            raise HTTPException(status_code=400, detail="Motivazione obbligatoria per ko/unsure")
        if reason.lower() == "altro" and not reason_text:
            raise HTTPException(status_code=400, detail="Per 'Altro' serve il testo della motivazione")
    return models.upsert_review(u["id"], target_key, judgement, reason, reason_text)


@app.post("/lab/api/comments/{target_key:path}", include_in_schema=False)
def api_add_comment(target_key: str, request: Request, payload: dict = Body(...)):
    u = require_user(request)
    body = (payload.get("body") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Testo del commento mancante")
    return models.create_comment(u["id"], target_key, body)


@app.put("/lab/api/comments/{comment_id}", include_in_schema=False)
def api_edit_comment(comment_id: int, request: Request, payload: dict = Body(...)):
    u = require_user(request)
    body = (payload.get("body") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Testo del commento mancante")
    ok = models.update_comment(comment_id, u["id"], body=body, is_admin=(u["role"] == "admin"))
    if not ok:
        raise HTTPException(status_code=403, detail="Non puoi modificare questo commento")
    return {"ok": True}


@app.delete("/lab/api/comments/{comment_id}", include_in_schema=False)
def api_delete_comment(comment_id: int, request: Request):
    u = require_user(request)
    ok = models.delete_comment(comment_id, u["id"], is_admin=(u["role"] == "admin"))
    if not ok:
        raise HTTPException(status_code=403, detail="Non puoi eliminare questo commento")
    return {"ok": True}


# ============================== admin API ==============================

@app.get("/lab/admin/stats", include_in_schema=False)
def admin_stats(request: Request):
    require_admin(request)
    return models.admin_stats()


@app.get("/lab/admin/users", include_in_schema=False)
def admin_list_users(request: Request):
    require_admin(request)
    return models.list_users()


@app.post("/lab/admin/users", include_in_schema=False)
def admin_create_user(request: Request, payload: dict = Body(...)):
    require_admin(request)
    username = (payload.get("username") or "").strip()
    password = payload.get("password") or ""
    role = payload.get("role") or "tester"
    if not username or not password:
        raise HTTPException(status_code=400, detail="username e password obbligatori")
    if role not in ("admin", "tester"):
        raise HTTPException(status_code=400, detail="role deve essere admin|tester")
    if models.get_user_by_username(username):
        raise HTTPException(status_code=400, detail="Username già esistente")
    uid = models.create_user(username, hash_password(password), role=role, must_change_password=True)
    return models.get_user(uid)


@app.put("/lab/admin/users/{user_id}", include_in_schema=False)
def admin_update_user(user_id: int, request: Request, payload: dict = Body(...)):
    me = require_admin(request)
    u = models.get_user(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    new_role = payload.get("role")
    new_active = payload.get("active")
    new_password = payload.get("password")
    # Evita di disattivare/declassare l'ultimo admin attivo
    if (new_role == "tester" or new_active is False) and u["role"] == "admin":
        admins = [x for x in models.list_users() if x["role"] == "admin" and x["active"]]
        if len(admins) == 1 and admins[0]["id"] == user_id:
            raise HTTPException(status_code=400, detail="Non puoi rimuovere l'unico admin attivo")
    # Se l'admin sta resettando la propria password, NON forzare must_change_password
    # (sta cambiando volontariamente la sua; obbligarlo a re-cambiarla è un loop).
    force_change = bool(new_password) and (user_id != me["id"])
    models.update_user(
        user_id,
        role=new_role,
        active=(None if new_active is None else bool(new_active)),
        password_hash=(hash_password(new_password) if new_password else None),
        must_change_password=(True if force_change else (False if new_password and user_id == me["id"] else None)),
    )
    return models.get_user(user_id)


@app.delete("/lab/admin/users/{user_id}", include_in_schema=False)
def admin_delete_user(user_id: int, request: Request):
    me = require_admin(request)
    if user_id == me["id"]:
        raise HTTPException(status_code=400, detail="Non puoi eliminare te stesso")
    u = models.get_user(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    # Protezione ultimo admin
    if u["role"] == "admin":
        admins = [x for x in models.list_users() if x["role"] == "admin" and x["active"]]
        if len(admins) <= 1:
            raise HTTPException(status_code=400, detail="Non puoi eliminare l'ultimo admin")
    models.delete_user(user_id)
    return {"ok": True}


@app.get("/lab/admin/cases", include_in_schema=False)
def admin_list_cases(request: Request):
    require_admin(request)
    return models.list_all_cases()


@app.get("/lab/admin/export/state.xlsx", include_in_schema=False)
def admin_export_state(request: Request, revised: str = "0"):
    """Export Excel multi-foglio dello stato dei test (solo admin).
    Parametro query ?revised=1 applica i messaggi rivisti all'azione primaria."""
    require_admin(request)
    _use_revised = str(revised) in ("1", "true", "True")
    import io as _io
    from datetime import datetime as _dt
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installato sul container")

    from tangible_lab.db import get_conn
    from nba_engine import generate_client_nba, generate_lead_nba
    from nba_api import load_dataset

    # ---- collect reviews/comments per target_key ----
    with get_conn() as conn:
        target_stats = {}
        for r in conn.execute("SELECT target_key, judgement FROM reviews"):
            t = target_stats.setdefault(r["target_key"], {"ok": 0, "ko": 0, "unsure": 0, "comments": 0, "last_review_at": ""})
            t[r["judgement"]] = t.get(r["judgement"], 0) + 1
        for r in conn.execute("SELECT target_key, MAX(updated_at) AS last FROM reviews GROUP BY target_key"):
            target_stats.setdefault(r["target_key"], {"ok": 0, "ko": 0, "unsure": 0, "comments": 0, "last_review_at": ""})["last_review_at"] = r["last"]
        for r in conn.execute("SELECT target_key, COUNT(*) AS n FROM comments GROUP BY target_key"):
            target_stats.setdefault(r["target_key"], {"ok": 0, "ko": 0, "unsure": 0, "comments": 0, "last_review_at": ""})["comments"] = r["n"]

    # ---- NBA enrich per ogni target_key (predef o case) ----
    dataset = load_dataset()
    clients_by_id = {c.get("client_id"): c for c in (dataset.get("clients") or [])}
    leads_by_id   = {l.get("lead_id"): l for l in (dataset.get("leads") or [])}

    def _nba_for_target(key: str):
        """Ritorna (kind, type, record_id, tier, score, strategy, primary_action, action_category) o None."""
        parts = key.split(":")
        if len(parts) < 2:
            return None
        kind = parts[0]
        rec_type = parts[1] if len(parts) > 2 else "?"
        rec_id = parts[-1]
        if kind == "predef":
            try:
                if rec_type == "client":
                    rec = clients_by_id.get(rec_id)
                    if not rec: return (kind, rec_type, rec_id, "", "", "", "", "")
                    out = generate_client_nba(rec)
                elif rec_type == "lead":
                    rec = leads_by_id.get(rec_id)
                    if not rec: return (kind, rec_type, rec_id, "", "", "", "", "")
                    out = generate_lead_nba(rec)
                else:
                    return (kind, rec_type, rec_id, "", "", "", "", "")
                if not out: return (kind, rec_type, rec_id, "", "", "", "", "")
                primary = next((a for a in (out.get("recommended_actions") or []) if a.get("primary")), None)
                return (kind, rec_type, rec_id,
                        out.get("priority_tier") or "",
                        round(out.get("priority_score") or 0, 1),
                        out.get("strategic_category") or "",
                        (primary.get("recommended_action") if primary else ""),
                        (primary.get("action_category") if primary else ""))
            except Exception:
                return (kind, rec_type, rec_id, "", "", "", "", "")
        elif kind == "case":
            return (kind, "case", rec_id, "", "", "", "", "")
        return (kind, rec_type, rec_id, "", "", "", "", "")

    # ---- build workbook ----
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F1C3D")
    head_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"

    # --- Sheet 1: Riepilogo ---
    ws = wb.active
    ws.title = "Riepilogo"
    stats = models.admin_stats()
    ws["A1"] = "NBA Studio — Stato dei test"; ws["A1"].font = Font(bold=True, size=14, color="1F1C3D")
    ws["A2"] = "Generato il"; ws["B2"] = _dt.utcnow().isoformat(timespec="seconds") + " UTC"
    ws.append([])
    ws.append(["KPI", "Valore"])
    style_row = lambda r: [ws.cell(row=r, column=c).__setattr__("font", bold) for c in (1,)]
    ws.append(["Utenti attivi", stats["users"]["active"]])
    ws.append(["Utenti totali", stats["users"]["total"]])
    ws.append(["Anagrafiche analizzate", len(target_stats)])
    ws.append(["Giudizi totali", stats["reviews"]["total"]])
    ws.append(["  Corretti", stats["reviews"]["by_judgement"].get("ok", 0)])
    ws.append(["  Sbagliati", stats["reviews"]["by_judgement"].get("ko", 0)])
    ws.append(["  Da verificare", stats["reviews"]["by_judgement"].get("unsure", 0)])
    ws.append(["Commenti totali", stats["comments"]["total"]])
    ws.append(["Casi salvati NBA", stats["cases"]["total"]])
    ws.append(["  di cui condivisi", stats["cases"]["shared"]])
    for c in ws["A4":"A14"]:
        c[0].font = bold
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.append([])
    ws.append(["Attività per utente"]); ws.cell(row=ws.max_row, column=1).font = bold
    user_table_start = ws.max_row + 1
    ws.append(["Utente", "Ruolo", "Attivo", "Giudizi", "Casi salvati", "Commenti"])
    for c in range(1, 7):
        ws.cell(row=user_table_start, column=c).font = head_font
        ws.cell(row=user_table_start, column=c).fill = head_fill
    for u in stats["per_user"]:
        ws.append([u["username"], u["role"], "sì" if u["active"] else "no",
                   u["reviews"], u["cases"], u["comments"]])

    # --- Sheet 2: Stato per anagrafica ---
    ws2 = wb.create_sheet("Stato per anagrafica")
    ws2.append([
        "target_key", "kind", "tipo", "record_id", "tier", "score", "strategy",
        "primary_action", "n_giudizi", "corretti", "sbagliati", "da_verificare",
        "motivazione", "n_commenti", "ultimo_giudizio"
    ])
    style_header(ws2, 15)
    # Mappa target→motivazione (single-user: una review per target; in multi prende la più recente)
    reason_by_target = {}
    for r in models.list_all_reviews_export():
        if r["target_key"] not in reason_by_target:
            txt = r.get("reason") or ""
            if txt and r.get("reason_text"):
                txt = f'{txt} — {r["reason_text"]}'
            reason_by_target[r["target_key"]] = txt
    rows_for_target = []
    for key, st in target_stats.items():
        nba = _nba_for_target(key) or ("", "", "", "", "", "", "", "")
        kind, rec_type, rec_id, tier, score, strategy, primary, act_cat = nba
        # Applica messaggi rivisti se richiesto dal parametro ?revised=1
        if _use_revised and primary:
            primary = lab_messages.revise(act_cat, primary)
        n_giudizi = st["ok"] + st["ko"] + st["unsure"]
        rows_for_target.append([
            key, kind, rec_type, rec_id, tier, score, strategy, primary,
            n_giudizi, st["ok"], st["ko"], st["unsure"],
            reason_by_target.get(key, ""),
            st["comments"], st["last_review_at"]
        ])
    # Ordina per priorità (tier CRITICAL prima, poi HIGH, ecc.)
    tier_rank = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "": 4}
    rows_for_target.sort(key=lambda r: (tier_rank.get(r[4], 5), -r[8]))
    for r in rows_for_target:
        ws2.append(r)
    # Larghezza colonne
    widths = [22, 8, 8, 10, 10, 8, 14, 60, 10, 9, 10, 13, 30, 11, 22]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Sheet 3: Giudizi ---
    ws3 = wb.create_sheet("Giudizi")
    ws3.append(["id", "utente", "target_key", "kind", "tipo", "record_id",
                "giudizio", "motivazione", "dettaglio", "creato", "aggiornato"])
    style_header(ws3, 11)
    for r in models.list_all_reviews_export():
        nba = _nba_for_target(r["target_key"]) or ("", "", "", "", "", "", "", "")
        kind, rec_type, rec_id, *_ = nba
        ws3.append([r["id"], r["username"], r["target_key"], kind, rec_type, rec_id,
                    r["judgement"], r.get("reason") or "", r.get("reason_text") or "",
                    r["created_at"], r["updated_at"]])
    for i, w in enumerate([7, 16, 24, 9, 9, 12, 13, 28, 30, 22, 22], start=1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Sheet 4: Note (commenti) ---
    ws4 = wb.create_sheet("Note")
    ws4.append(["id", "utente", "target_key", "kind", "tipo", "record_id",
                "nota", "creato", "aggiornato"])
    style_header(ws4, 9)
    for r in models.list_all_comments_export():
        nba = _nba_for_target(r["target_key"]) or ("", "", "", "", "", "", "", "")
        kind, rec_type, rec_id, *_ = nba
        ws4.append([r["id"], r["username"], r["target_key"], kind, rec_type, rec_id,
                    r["body"], r["created_at"], r["updated_at"]])
    for i, w in enumerate([7, 16, 24, 9, 9, 12, 70, 22, 22], start=1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"stato-test_{_dt.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/lab/admin/reviews", include_in_schema=False)
def admin_list_reviews(request: Request, format: Optional[str] = None):
    require_admin(request)
    rows = models.list_all_reviews_export()
    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["id", "target_key", "user", "judgement", "created_at", "updated_at"])
        for r in rows:
            w.writerow([r["id"], r["target_key"], r["username"], r["judgement"], r["created_at"], r["updated_at"]])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="reviews.csv"'},
        )
    return rows


# ============================== Check-up Vittoria ==============================

@app.get("/lab/api/checkup/data", include_in_schema=False)
def api_checkup_data(request: Request):
    require_user(request)
    return checkup_engine.load_data()


@app.post("/lab/api/checkup/compute", include_in_schema=False)
def api_checkup_compute(request: Request, payload: dict = Body(...)):
    require_user(request)
    answers = payload.get("answers") or {}
    errors = checkup_engine.validate_answers(answers)
    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors})
    return checkup_engine.compute(answers)


@app.get("/lab/api/checkup/cases", include_in_schema=False)
def api_checkup_list(request: Request):
    u = require_user(request)
    return checkup_models.list_visible_to(u["id"])


@app.post("/lab/api/checkup/cases", include_in_schema=False)
def api_checkup_create(request: Request, payload: dict = Body(...)):
    u = require_user(request)
    name = (payload.get("name") or "").strip() or "(senza nome)"
    answers = payload.get("answers") or {}
    shared = bool(payload.get("shared", False))
    notes = payload.get("notes") or ""
    errors = checkup_engine.validate_answers(answers)
    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors})
    result = checkup_engine.compute(answers)
    cid = checkup_models.create(u["id"], name, answers, result, shared=shared, notes=notes)
    return checkup_models.get(cid)


@app.put("/lab/api/checkup/cases/{case_id}", include_in_schema=False)
def api_checkup_update(case_id: int, request: Request, payload: dict = Body(...)):
    u = require_user(request)
    c = checkup_models.get(case_id)
    if not c:
        raise HTTPException(status_code=404, detail="Caso non trovato")
    if c["owner_id"] != u["id"] and u["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo l'autore (o admin) può modificare")
    answers = payload.get("answers")
    result = None
    if answers is not None:
        errors = checkup_engine.validate_answers(answers)
        if errors:
            raise HTTPException(status_code=400, detail={"errors": errors})
        result = checkup_engine.compute(answers)
    checkup_models.update(
        case_id,
        name=payload.get("name"),
        answers=answers,
        result=result,
        shared=payload.get("shared"),
        notes=payload.get("notes"),
    )
    return checkup_models.get(case_id)


@app.post("/lab/admin/checkup/reseed", include_in_schema=False)
def admin_checkup_reseed(request: Request):
    """Re-popola gli scenari demo del Check-up. Idempotent: salta scenari con nome già esistente."""
    me = require_admin(request)
    import json as _json
    seed_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "checkup", "seed_cases.json")
    if not os.path.exists(seed_path):
        raise HTTPException(status_code=500, detail="seed file mancante")
    scenarios = _json.loads(open(seed_path, "r", encoding="utf-8").read())
    existing_names = {c["name"] for c in checkup_models.list_all()}
    created = 0
    skipped = 0
    for s in scenarios:
        if s["name"] in existing_names:
            skipped += 1
            continue
        answers = s.get("answers") or {}
        errs = checkup_engine.validate_answers(answers)
        if errs:
            skipped += 1
            continue
        result = checkup_engine.compute(answers)
        checkup_models.create(me["id"], s["name"], answers, result, shared=True, notes="Scenario demo")
        created += 1
    return {"created": created, "skipped": skipped, "total": len(scenarios)}


@app.delete("/lab/api/checkup/cases/{case_id}", include_in_schema=False)
def api_checkup_delete(case_id: int, request: Request):
    u = require_user(request)
    c = checkup_models.get(case_id)
    if not c:
        raise HTTPException(status_code=404, detail="Caso non trovato")
    if c["owner_id"] != u["id"] and u["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo l'autore (o admin) può eliminare")
    checkup_models.delete(case_id)
    return {"ok": True}


# ============================== breakdown punteggio (Tangible) ==============================

def _flatten_schema_cfg(schema_cfg):
    out = {}
    for section, content in (schema_cfg or {}).items():
        if isinstance(content, dict):
            sec = {}
            for k, node in content.items():
                if isinstance(node, dict) and "value" in node:
                    sec[k] = node["value"]
                else:
                    sec[k] = node
            out[section] = sec
        else:
            out[section] = content
    return out


def _clamp01_100(v):
    return max(0.0, min(100.0, float(v)))


class _NullCtx:
    def __enter__(self): return self
    def __exit__(self, *a): return False


@app.post("/lab/breakdown/client", include_in_schema=False)
def lab_breakdown_client(payload: dict = Body(...)):
    raw = payload.get("client")
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="Body must include {client:<obj>, config?:<obj>}")
    cfg_schema = payload.get("config")
    cfg = _flatten_schema_cfg(cfg_schema) if cfg_schema else get_config()
    ctx = temporary_config(cfg) if cfg_schema else _NullCtx()
    with ctx:
        c = client_from_dict(raw)
        triggers = detect_client_triggers(c, cfg)
        w = cfg["client_weights"]
        urgency = client_urgency_score(triggers, cfg)
        value, _ = client_value_score(c, triggers, cfg, debug=False)
        opportunity = client_opportunity_score(c, triggers, cfg)
        recency = client_recency_score(c)
        wu = float(w["urgency"]); wv = float(w["value"])
        wo = float(w["opportunity"]); wr = float(w["recency"])
        contributions = [
            {"factor": "urgency",     "score": _clamp01_100(urgency),     "weight": wu, "contribution": _clamp01_100(urgency)     * wu},
            {"factor": "value",       "score": _clamp01_100(value),       "weight": wv, "contribution": _clamp01_100(value)       * wv},
            {"factor": "opportunity", "score": _clamp01_100(opportunity), "weight": wo, "contribution": _clamp01_100(opportunity) * wo},
            {"factor": "recency",     "score": _clamp01_100(recency),     "weight": wr, "contribution": _clamp01_100(recency)     * wr},
        ]
        base = sum(item["contribution"] for item in contributions)
        bonus_rel = float(client_relationship_modifier(c))
        before_clamp = base + bonus_rel
        clamped = max(0.0, min(100.0, before_clamp))
        churn_or = bool(churn_override(triggers, cfg))
        crit_threshold = float(cfg["tiers"]["CRITICAL"])
        final = max(clamped, crit_threshold) if churn_or else clamped
        return {
            "type": "client",
            "contributions": contributions,
            "base_score": base,
            "bonuses": [
                {"label": "Modificatore relazione", "value": bonus_rel},
            ],
            "before_clamp": before_clamp,
            "clamped": clamped,
            "churn_override_applied": churn_or,
            "critical_threshold": crit_threshold,
            "final_score": final,
        }


@app.post("/lab/breakdown/lead", include_in_schema=False)
def lab_breakdown_lead(payload: dict = Body(...)):
    raw = payload.get("lead")
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="Body must include {lead:<obj>, config?:<obj>}")
    cfg_schema = payload.get("config")
    cfg = _flatten_schema_cfg(cfg_schema) if cfg_schema else get_config()
    ctx = temporary_config(cfg) if cfg_schema else _NullCtx()
    with ctx:
        l = lead_from_dict(raw)
        triggers = detect_lead_triggers(l, cfg)
        if "COVERAGE_START_SOON" in triggers:
            d = int(triggers["COVERAGE_START_SOON"]["days_remaining"])
            urg = 85.0 if d <= 3 else 75.0 if d <= 7 else 60.0
        elif "NEW_LEAD" in triggers:
            h = int(triggers["NEW_LEAD"]["hours_since_creation"])
            urg = 90.0 if h <= 4 else 80.0 if h <= 12 else 70.0 if h <= 24 else 55.0
        elif "STALE_LEAD" in triggers:
            urg = 50.0
        else:
            urg = 30.0
        if "HIGH_VALUE_QUOTE" in triggers:
            val = 100.0
        elif "QUOTE_READY" in triggers:
            p = float(triggers["QUOTE_READY"]["premium"])
            val = 80.0 if p >= 400 else 60.0 if p >= 300 else 40.0
        else:
            val = 20.0
        d = l.last_contact_days
        if d is None:
            tim = 90.0
        elif int(d) == 0:
            tim = 30.0
        elif int(d) == 1:
            tim = 50.0
        elif 2 <= int(d) <= 3:
            tim = 70.0
        else:
            tim = 85.0
        w = cfg["lead_weights"]
        wu = float(w["urgency"]); wv = float(w["value"]); wt = float(w["timing"])
        contributions = [
            {"factor": "urgency", "score": urg, "weight": wu, "contribution": urg * wu},
            {"factor": "value",   "score": val, "weight": wv, "contribution": val * wv},
            {"factor": "timing",  "score": tim, "weight": wt, "contribution": tim * wt},
        ]
        base = sum(item["contribution"] for item in contributions)
        # Il motore somma un bonus proporzionale al numero di trigger (cap a 24).
        bonus_trigger = min(len(triggers) * 6.0, 24.0)
        before_clamp = base + bonus_trigger
        clamped = max(0.0, min(100.0, before_clamp))
        return {
            "type": "lead",
            "contributions": contributions,
            "base_score": base,
            "bonuses": [
                {"label": "Bonus trigger", "value": bonus_trigger},
            ],
            "before_clamp": before_clamp,
            "clamped": clamped,
            "churn_override_applied": False,
            "critical_threshold": float(cfg["tiers"]["CRITICAL"]),
            "final_score": clamped,
        }


@app.post("/lab/admin/dataset/import", include_in_schema=False)
def admin_import_dataset(request: Request, payload: dict = Body(...)):
    """Importa il dataset reale in locale (solo admin). Sostituisce quello corrente.

    Il dato non sta mai nel repo/exe: viene scritto in DATASET_PATH
    (%APPDATA%\\NBAStudio\\dataset.json) e la cache del backend viene azzerata.
    """
    require_admin(request)
    import nba_api as _na

    clients = payload.get("clients")
    leads = payload.get("leads")
    if not isinstance(clients, list) or not isinstance(leads, list):
        raise HTTPException(
            status_code=400,
            detail="Formato non valido: il file deve contenere le liste 'clients' e 'leads'.",
        )

    target = _na.DATASET_PATH
    os.makedirs(os.path.dirname(target) or ".", exist_ok=True)
    tmp = target + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    os.replace(tmp, target)   # scrittura atomica

    with _na._DATA_LOCK:       # reload a caldo lock-safe (come nba_api.dataset_reload)
        _na._DATA = None
    return {"clients": len(clients), "leads": len(leads)}


@app.post("/lab/admin/config/reset", include_in_schema=False)
def admin_config_reset(request: Request):
    """Ripristina la config di default del cliente (re-seed) e ricarica il motore."""
    require_admin(request)
    import nba_config as _nc
    src = os.path.join(_REPO_ROOT, "nba_config.json")
    if not os.path.exists(src):
        raise HTTPException(status_code=500, detail="Config di default non trovata")
    shutil.copy2(src, _nc.CONFIG_JSON_PATH)
    _nc.reload_config()
    return {"status": "ok"}


@app.get("/lab/api/messages-map", include_in_schema=False)
def lab_messages_map():
    """Mappa dei messaggi rivisti per il frontend (toggle 'Messaggi rivisti')."""
    import json as _json
    with open(lab_messages.MAP_PATH, encoding="utf-8") as f:
        return _json.load(f)


# ============================== static mount (alla FINE, così non intercetta /lab/api/*) ==============================

app.mount("/lab", StaticFiles(directory=LAB_DIR, html=True), name="tangible_lab")


# ============================== launcher ==============================

def _find_free_port(host: str) -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((host, 0))
        return int(s.getsockname()[1])


def _open_browser(url: str) -> None:
    try:
        webbrowser.open(url)
    except Exception:
        pass


def run(host: str = "127.0.0.1", port: int = 8000, open_ui: bool = True) -> None:
    import uvicorn
    chosen_port = port
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.2)
            if s.connect_ex((host, port)) == 0:
                chosen_port = _find_free_port(host)
    except Exception:
        chosen_port = port
    url = f"http://{host}:{chosen_port}/lab/"
    if open_ui:
        threading.Timer(0.6, _open_browser, args=(url,)).start()
    print(f"[tangible_lab] NBA Studio on {url}")
    uvicorn.run(app, host=host, port=chosen_port)


if __name__ == "__main__":
    run()
